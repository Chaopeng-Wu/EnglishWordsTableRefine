# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import string

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook

# 读取源表
original_wb = openpyxl.load_workbook("研4500.xlsx", read_only=True)
# 创建新表
new_wb = Workbook()
# 移除新表的原始页
new_wb.remove(new_wb['Sheet'])
# sheets = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
# 'V', 'W', 'Y', 'Z']
# 为新表规定名称
sheets = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']
# 用以储存每一页的名字以及其对应的词条数
Range = []
# 为新表创建所有页
for sheetName in sheets:
    # 将源表的页提出
    original_ws: Worksheet = original_wb[sheetName]
    # 为新表的页命名
    new_ws = new_wb.create_sheet(sheetName)
    # 用于储存源页的内容
    contents: list[str] = []
    # 获取源页的内容
    for row in original_ws.rows:
        for cell in row:
            cellContent = cell.value.strip(string.digits)
            cellContent = cellContent.strip(' ')
            contents.append(cellContent)

    # 储存源页的内容中的每一项， 一般 group[0] 为第一组词，group[0][0]为单词 group[1][1] 为翻译
    group = []

    for index in range(len(contents)):
        group.append(contents[index].split(' ', 1))

    # 遍历 group，为每一项去除空字符
    for y in range(len(group)):
        for x in range(2):
            if x == 1:
                if len(group[y]) == 1:
                    continue
            # 去除空字符
            group[y][x] = group[y][x].strip(' ')

    cache = 1

    # 按照格式写入
    for row in range(len(group)):
        for column in range(2):
            if column == 1:
                if len(group[row]) == 1:
                    continue
            # 每三行空一行
            if (row + cache) % 4 == 0:
                if column == 0:
                    # 记录下空出来的行数
                    cache += 1
                    # 为空白行写入空字符，占个位置
                    new_ws.cell(row + cache, column + 1).value = " "
            new_ws.cell(row + cache, column + 1).value = group[row][column]
    # 在第一行写入当前页的项总数
    new_ws.insert_rows(1, 1)
    new_ws.cell(1, 1).value = len(group)
    # 储存每页的词条数
    item = [sheetName, len(group)]
    Range.append(item)

new_wb.save('4500.xlsx')
# 排序 key = lambda 的意思是比大小， v 是给参数，代表数组中的每个元素
Range.sort(key=lambda v: v[1])
for item in range(len(Range)):
    print(Range[item][0], ":", Range[item][1])
